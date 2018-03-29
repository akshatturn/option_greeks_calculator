import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook

def getvalue(element):
  elems = soup.select(element)
  return elems[0].text.strip()

driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver')
driver.get('https://www.nseindia.com/live_market/dynaContent/live_watch/option_chain/optionKeys.jsp?symbolCode=-10002&symbol=NIFTY&symbol=NIFTY&instrument=-&date=-&segmentLink=17&symbolCount=2&segmentLink=17')

html = driver.page_source
soup = BeautifulSoup(html, "lxml")
now = datetime.now()

wb = Workbook()
ws = wb.active

# Timestamp
v = getvalue('#niftyDiv > p.right > font > nobr')
ws['A1']= 'DATE : '
ws['B1']= '%s:%s' % (now.hour,now.minute)
ws['C1']= 'QUOTE :'
ws['D1']= '%s' % str(v)

# Insert Thread
list_for_thread = []
for x in range(2,23):
  t = '#octable > thead > tr:nth-of-type(2) > th:nth-of-type(%s)' % x
  list_for_thread.append(getvalue(t))
ws.append(list_for_thread)

# Insert Data

try:
  for x in range(1,200):
    list_row = []
    for y in range(2,23):
       s = '#octable > tbody > tr:nth-of-type(%s) > td:nth-of-type(%s)' % (x,y)
       list_row.append(getvalue(s))
    ws.append(list_row)
except Exception:
# Save file
  p = "/home/akshat/Desktop/Practice/Nifty50(%s-%s-%s).xlsx" %(str(now.day),str(now.month),str(now.year))
  wb.save(p)

time.sleep(5) 
driver.quit()
