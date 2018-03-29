import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook


def getvalue(element):
  elems = driver.find_elements_by_css_selector(element)
  return elems[0].text.strip()

#Enter Values 
def enter_values(spot,strike,volatility,interest):
  driver.find_element_by_css_selector('#input-spot').send_keys(Keys.CONTROL + 'a' + Keys.NULL,spot)
  driver.find_element_by_css_selector('#input-strike').send_keys(Keys.CONTROL + 'a' + Keys.NULL,strike)
  driver.find_element_by_css_selector('#input-volt').send_keys(Keys.CONTROL + 'a' + Keys.NULL,volatility)
  driver.find_element_by_css_selector('#input-intrate').send_keys(Keys.CONTROL + 'a' + Keys.NULL,interest)
  driver.find_element_by_css_selector('#datetimepicker').send_keys(Keys.CONTROL + 'a' + Keys.NULL,'2','0','1','7','1','2','2','8')  
  driver.find_element_by_css_selector('#calc-button').click()

#File name
now = datetime.now()
filename = "Nifty50(%s-%s-%s).xlsx" %(str(now.day),str(now.month),str(now.year))

wb1 = load_workbook(filename)
ws1 = wb1.active
spott = ws1['D1'].value
ws2 = wb1.create_sheet("Greeks")

ws2.append(['CEPremium','CE Delt','CE Gamma','CE Theta','CE Vega','CE RHo','CE LTP','Strike','PE LTP','PE Delta','PE Gamma','PE Theta','PE Vega','PE Rho','PEPremium'])

#Fetch from sheet and put into sheet
def fetch_sheet(rw):
  strik = ws1.cell(row=rw,column = 11).value
  ivce = ws1.cell(row=rw,column = 4).value
  if ivce =='-':
    ivce = 0
  ivpe = ws1.cell(row=rw,column = 18).value
  if ivpe =='-':
    ivpe = 0
  celtp = ws1.cell(row=rw,column = 5).value
  peltp = ws1.cell(row=rw,column = 17).value
  l =[]
  enter_values(spott,strik,ivce,10)
  l.append(getvalue('#call-option-prem-value'))
  l.append(getvalue('#call-option-delta-value'))
  l.append(getvalue('#option-gamma-value'))
  l.append(getvalue('#call-option-theta-value'))
  l.append(getvalue('#option-vega-value'))
  l.append(getvalue('#class-option-rho-value'))
  l.append(celtp)
  l.append(strik)
  enter_values(spott,strik,ivpe,10)
  l.append(peltp)
  l.append(getvalue('#put-option-delta-value'))
  l.append(getvalue('#option-gamma-value'))
  l.append(getvalue('#put-option-theta-value'))
  l.append(getvalue('#option-vega-value'))
  l.append(getvalue('#put-option-rho-value'))
  l.append(getvalue('#put-option-prem-value'))
  ws2.append(l)


# Open Option calculator
driver = webdriver.Chrome('/usr/lib/chromium-browser/chromedriver')
driver.get('https://zerodha.com/tools/black-scholes/')


html = driver.page_source
soup = BeautifulSoup(html, "lxml")

try:
  for i in range(3,201):
    fetch_sheet(i)
except Exception:
  print "Completed!!"
  wb1.save(filename)

time.sleep(3) 
driver.quit()
